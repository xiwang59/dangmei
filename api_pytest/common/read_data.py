#-*-coding:utf-8-*-
import self as self
from openpyxl import load_workbook
from api_pytest.common import login_token

#获取所有的excel测试用例的数据
class ReadData:
    def __init__(self,path,sheetname1,sheetname2):
        self.path=path
        self.sheetname1=sheetname1
        self.sheetname2 = sheetname2
        self.wb=load_workbook(self.path)  #打开工作簿
        self.sheet1 = self.wb["%s"%self.sheetname1]   # #定位到你所需操作的目标表单，也就是sheet
        self.sheet2 = self.wb["%s"%self.sheetname2]   # #定位到你所需操作的目标表单，也就是sheet


    #获取初始化数据
    def ReadIint(self):
        row = self.sheet2.max_row  # 获取最大的行
        col = self.sheet2.max_column  # 获取最大的列
        init_data={}
        for j in (2,row-1):
            key=self.sheet2.cell(j,1).value
            init_data[key]=self.sheet2.cell(j,2).value
        init_data["${nummer}"]=str(int(init_data["${nummer}"]))
        return init_data

    #获取所有的数据
    def Read(self):
        initdata = self.ReadIint()   #实例化初始化的数据
        # print(initdata)
        row = self.sheet1.max_row  # 获取最大的行
        col = self.sheet1.max_column  # 获取最大的列
        l=[]
        for i in range(row-1):
            d={}
            #获取每行数据，用字典的形式存储
            for j in range(col):
                keydata = self.sheet1.cell(1,j+1).value
                result=self.sheet1.cell(i+2,j+1).value
                d[keydata]=result


            if d['request_data'] is not None:  #判断下请求的数据是否为空，因为find函数是对字符串进行查找
                for key,value in initdata.items():  #获取初始值的key，value
                    if d['request_data'].find(key) != -1:   #查找key是否在 request_data中  find函数如果不在其中，则显示-1
                        new_data=d['request_data'].replace(key,value)   #将key替换成value
                        d['request_data']=new_data


            #将当次获取的token替换成最新的
            new_token=d['request_header'].replace("${token}",login_token.token)
            d['request_header']=new_token
            # print(d['request_header'])

            l.append(d)
            # print(l)
        return l
    # 修改初始化数据
    def updata_init_data(self):
        new_init_data=self.sheet2.cell(2,2).value=str(int(self.sheet2.cell(2,2).value)+1)
        self.save_excel()
        # print(new_init_data)


        #保存excel
    def save_excel(self):
        self.wb.save(self.path)

if __name__ == '__main__':
    p = "D:/PycharmProjects/again/api_request/testdata/dangmei_api.xlsx"
    sj="用例"
    s="初始值"
    # a=ReadData(p,sj,s).ReadIint()
    b=ReadData(p,sj,s).Read()
    # c=ReadData(p,sj,s).updata_init_data()
    # print(a)
    print(b)
    # print(c)




